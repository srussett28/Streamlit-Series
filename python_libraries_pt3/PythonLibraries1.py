
import streamlit as st
import os

import pandas as pd
import numpy as np


import matplotlib.pyplot as plt
import seaborn as sns
import plotly.express as px


from sklearn.linear_model import LinearRegression
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_squared_error


from sqlalchemy import create_engine # type: ignore


import requests

st.title("Essential Python Libraries for Analysts, Scientists, and Engineers")
image_path_python = "pythoncloud.jpg"
st.image(image_path_python, use_container_width=True)
st.write("Welcome to the Essential Python Libraries for Analysts, Scientists, and Engineers! This is a Streamlit app that showcases the most important Python libraries for data analysis, visualization, and machine learning. The libraries that will be covered in this app are: Pandas, NumPy, Matplotlib, Seaborn, Plotly, and Scikit-Learn. Each library will have its own section where we will explore the most important features and functions. Let's get started!")


st.title("Essential Libaries for Data Analysts (If your up for it)")
image_path_analyst = "data-analyst1.jpg"
st.image(image_path_analyst, use_container_width=True)


st.subheader("1. Pandas: The Data Manipulation Powerhouse")
image_path_pandas = "pandas2.png"
st.image(image_path_pandas, use_container_width=True)
st.warning(""" Use Case: Sales Data Analysis (aka ‘Make It Pretty for the Executive Team’)""")
st.markdown(""" You’ve got sales data that looks like it was written in a hurry by a caffeinated jack russell. Pandas helps you clean it up, calculate revenues, and make it look like you spent hours on it (when it really took 15 minutes).
""")


st.markdown("""
- **`read_csv()`**: Reads your messy CSV file and pretends there is clean data inside.""")
st.code("""
import pandas as pd

# Read a CSV file
data = pd.read_csv("sales_data.csv")
print(data.head())
""", language="python")

st.markdown("""
- **`groupby()`**: Groups data like a champ.""")
st.code("""
# Group data by a column and calculate sum
grouped_data = data.groupby("Product")["Revenue"].sum()
print(grouped_data)
""", language="python")


st.markdown("""
- **`merge()`**: Matches data like a Tinder date—sometimes perfect, sometimes… eh.
""")
st.code("""
# Merge two DataFrames on a common column
df1 = pd.DataFrame({"ID": [1, 2], "Name": ["Alice", "Bob"]})
df2 = pd.DataFrame({"ID": [1, 2], "Score": [85, 90]})
merged_data = pd.merge(df1, df2, on="ID")
print(merged_data)
""", language="python")


st.markdown("""
- **`fillna()`**: Fills in the blanks.""")  

st.code("""
# Fill missing values with a default value
data["Revenue"] = data["Revenue"].fillna(0)
print(data.head())
""", language="python")


st.markdown("""
- **`describe()`**: Summarizes your data so you can sound smarter than you feel.""") 

st.code("""
# Get summary statistics
summary = data.describe()
print(summary)
""", language="python")



st.markdown("""
- **`pivot_table()`**: Creates fancy pivot tables that would make Excel jealous.""") 

st.code("""
# Create a pivot table
pivot = data.pivot_table(values="Revenue", index="Product", columns="Region", aggfunc="sum")
print(pivot)
""", language="python")


st.markdown("""
- **`dropna()`**: Kicks out rows or columns with missing data, like a bouncer at a nightclub.""")

st.code("""
# Drop rows with missing values
cleaned_data = data.dropna()
print(cleaned_data.head())
""", language="python")


st.markdown("""
- **`sort_values()`**: Sorts your data by column values because who doesn’t like order?""")

st.code("""
# Sort data by the 'Revenue' column in descending order
sorted_data = data.sort_values(by="Revenue", ascending=False)
print(sorted_data.head())
""", language="python")

st.markdown("""
- **`iloc[]` and `loc[]`**: Helps you find that one needle in the data haystack.""")

st.code("""
# iloc example: Get the first 5 rows
subset_iloc = data.iloc[:5]

# loc example: Get rows where 'Revenue' > 100
subset_loc = data.loc[data["Revenue"] > 100]
print(subset_iloc)
print(subset_loc)
""", language="python")


st.markdown("""
- **`apply()`**: Lets you apply your own logic to rows or columns—perfect for when Pandas doesn’t already do it for you.""")

st.code("""
# Apply a custom function to a column
data["Discounted_Price"] = data["Price"].apply(lambda x: x * 0.9)
print(data.head())
""", language="python")


st.markdown("""
- **`astype()`**: Converts data types, because sometimes your numbers come in disguised as text.""")

st.code("""
# Convert a column to float
data["Revenue"] = data["Revenue"].astype(float)
print(data.dtypes)
""", language="python")


st.markdown("""
- **`value_counts()`**: Counts how many times each value appears in a column, so you can spot trends or weird anomalies.""")

st.code("""
# Count the occurrences of each product
product_counts = data["Product"].value_counts()
print(product_counts)
""", language="python")

st.markdown("""
- **`concat()`**: Sticks multiple DataFrames together like a glue for data.""")

st.code("""
# Concatenate two DataFrames
df1 = pd.DataFrame({"ID": [1, 2], "Name": ["Alice", "Bob"]})
df2 = pd.DataFrame({"ID": [3, 4], "Name": ["Charlie", "Diana"]})
combined_data = pd.concat([df1, df2])
print(combined_data)
""", language="python")


st.markdown("""
- **`drop_duplicates()`**: Removes duplicate rows because redundancy in this context sucks.Why? Duplicate rows in your dataset can skew analysis, inflate metrics, and lead to incorrect insights. """)

st.code("""
# Remove duplicate rows
unique_data = data.drop_duplicates()
print(unique_data)
""", language="python")


st.markdown("""
- **`isnull()`**: Detects missing data faster than your boss finds typos in your reporting.""")

st.code("""
# Check for missing values
missing_data = data.isnull()
print(missing_data)
""", language="python")


st.markdown("""
- **`to_excel()`**: Exports your cleaned-up masterpiece back to Excel for the spreadsheet fans in your office.
""")


st.code("""
# Save the DataFrame to an Excel file
data.to_excel("cleaned_data.xlsx", index=False)
print("Data exported to Excel!")
""", language="python")


st.subheader(" NumPy: Number Cruncher")
image_path_pandas = "Numpy.png"
st.image(image_path_pandas,caption="https://github.com/numpy/numpy.org/issues/37", use_container_width=True)


st.markdown("""
- **`array()`**: Creates arrays that don’t complain like lists.  
  *(An array is a powerful, grid-like data structure provided by NumPy. Why it doesn’t complain: Arrays offer uniformity, speed, and efficient math operations!)*
""")

st.code("""
import numpy as np

# Create an array
arr = np.array([1, 2, 3, 4])
print(arr)
""", language="python")

st.markdown("""
- **`mean()`**: Finds the average, because math.
""")
st.code("""
# Calculate the mean
mean_value = np.mean(arr)
print(mean_value)
""", language="python")

st.markdown("""
- **`sum()`**: Adds things up faster than a calculator.
""")
st.code("""
# Compute the sum
total = np.sum(arr)
print(total)
""", language="python")

st.markdown("""
- **`reshape()`**: Reshapes arrays like a pro.
""")
st.code("""
# Reshape a 1D array into a 2D array
reshaped = arr.reshape(2, 2)
print(reshaped)
""", language="python")

st.markdown("""
- **`random()`**: Generates random numbers because why not(Be Careful, and Intelligent)?
""")
st.code("""
# Generate random numbers
random_numbers = np.random.rand(3)
print(random_numbers)
""", language="python")


st.subheader("Matplotlib: Paint Data Like Picasso")
image_path_matplotlib = "matplot3.png"
st.image(image_path_matplotlib, use_container_width=True)


st.markdown("""
- **`plot()`**: Connects the dots with lines like an old-school Etch A Sketch.  
  *(Use this for simple line charts, great for trends or continuous data.)*
""")
st.code("""
import matplotlib.pyplot as plt

# Create a line plot
x = [1, 2, 3, 4]
y = [10, 20, 25, 30]
plt.plot(x, y)
plt.title("Line Plot Example")
plt.show()
""", language="python")

st.markdown("""
- **`scatter()`**: Puts dots on a graph, perfect for showing relationships.  
  *(Use this for visualizing correlations or clusters.)*
""")
st.code("""
# Create a scatter plot
plt.scatter(x, y)
plt.title("Scatter Plot Example")
plt.show()
""", language="python")

image_path_graph = "Matplotgraph.png"
st.image(image_path_graph,caption="https://coding-blocks.github.io/DS-NOTES/1.2%20Introduction%20to%20Matplotlib.html", use_container_width=True)




st.subheader("Seaborn: Stylish Data Visualizations")
image_path_seaborn = "seaborn2.png"
st.image(image_path_seaborn,caption= "https://datascientest.com/en/seaborn_and_data_visualization" , use_container_width=True)


st.markdown("""
- **`heatmap()`**: Makes your data hot (or cold).  
  *(Use this to visualize correlations, null values, or matrix data.)*
""")
st.code("""
import seaborn as sns
import numpy as np

# Create a heatmap
data = np.random.rand(5, 5)
sns.heatmap(data, annot=True)
plt.title("Heatmap Example")
plt.show()
""", language="python")

st.markdown("""
- **`boxplot()`**: Puts your data in a box (literally) to show ranges and outliers.  
  *(Great for showing distributions and spotting anomalies.)*
""")
st.code("""
# Create a boxplot
sns.boxplot(data=[7, 8, 8, 9, 10, 10, 11])
plt.title("Boxplot Example")
plt.show()
""", language="python")

image_path_seaborngraph = "seaborngraph.png"
st.image(
    image_path_seaborngraph, 
    caption="[Seaborn Function Overview](https://seaborn.pydata.org/tutorial/function_overview.html)", 
    use_container_width=True
)


st.subheader("openpyxl: The Excel Whisperer")
image_path_openpyxl = "yxl.png"
st.image(image_path_openpyxl, use_container_width=True)


st.markdown("""
- **`load_workbook()`**: Opens an Excel file, no double-clicking required.  
  *(Use this to edit existing spreadsheets programmatically.)*
""")
st.code("""
from openpyxl import load_workbook

# Load an existing Excel workbook
workbook = load_workbook("example.xlsx")
sheet = workbook.active
print(sheet.title)
""", language="python")

st.markdown("""
- **`create_sheet()`**: Adds a shiny new sheet to your workbook.  
  *(Perfect for organizing data into multiple tabs.)*
""")
st.code("""
# Create a new sheet
workbook.create_sheet(title="NewSheet")
workbook.save("example.xlsx")
""", language="python")



st.title("Essential Libaries for Data Scientists")
image_path_science = "DS.png"
st.image(image_path_science, use_container_width=True)
st.warning("To avoid redundency, Pandas and Numpy are not listed here.The majority of uses won't deviate from the analyst write up. But if you have a special or more advanced use, send it my way.")

st.subheader("scikit-learn: A Machine Learning Maestro")
image_path_sklearn = "scikit2.png"
st.image(image_path_sklearn, use_container_width=True)


st.markdown("""
- **`train_test_split()`**: Splits data into training and testing sets.  
  *(Helps evaluate model performance effectively.)*
""")
st.code("""
from sklearn.model_selection import train_test_split

X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
""", language="python")


st.markdown("""
- **`LinearRegression()`**: Builds and evaluates linear regression models.  
  *(A classic choice for continuous data prediction.)*
""")
st.code("""
from sklearn.linear_model import LinearRegression

model = LinearRegression()
model.fit(X_train, y_train)
print(model.coef_, model.intercept_)
""", language="python")


st.markdown("""
- **`DecisionTreeRegressor()`**: Creates regression trees for flexible, non-linear relationships.  
  *(Great for capturing complex patterns in the data.)*
""")
st.code("""
from sklearn.tree import DecisionTreeRegressor

# Decision tree regression example
tree_model = DecisionTreeRegressor(max_depth=5, random_state=42)
tree_model.fit(X_train, y_train)
predictions = tree_model.predict(X_test)
""", language="python")


st.markdown("""
- **`RandomForestRegressor()`**: Builds an ensemble of decision trees for robust predictions.  
  *(Reduces overfitting and improves accuracy.)*
""")
st.code("""
from sklearn.ensemble import RandomForestRegressor

# Random forest regression example
forest_model = RandomForestRegressor(n_estimators=100, random_state=42)
forest_model.fit(X_train, y_train)
predictions = forest_model.predict(X_test)
""", language="python")


st.markdown("""
- **`Ridge()`**: Implements linear regression with L2 regularization.  
  *(Prevents overfitting by penalizing large coefficients.)*
""")
st.code("""
from sklearn.linear_model import Ridge

# Ridge regression example
ridge_model = Ridge(alpha=1.0)
ridge_model.fit(X_train, y_train)
predictions = ridge_model.predict(X_test)
""", language="python")


st.markdown("""
- **`Lasso()`**: Implements linear regression with L1 regularization.  
  *(Encourages sparsity by shrinking some coefficients to zero.)*
""")
st.code("""
from sklearn.linear_model import Lasso

# Lasso regression example
lasso_model = Lasso(alpha=0.1)
lasso_model.fit(X_train, y_train)
predictions = lasso_model.predict(X_test)
""", language="python")


st.markdown("""
- **`StandardScaler()`**: Standardizes features by removing the mean and scaling to unit variance.  
  *(Essential for algorithms sensitive to feature scaling.)*
""")
st.code("""
from sklearn.preprocessing import StandardScaler
scaler = StandardScaler()
X_scaled = scaler.fit_transform(X)
""", language="python")


st.markdown("""
- **`KMeans()`**: Clusters data into K distinct groups.  
  *(Useful for unsupervised learning tasks.)*
""")
st.code("""
from sklearn.cluster import KMeans

kmeans = KMeans(n_clusters=3, random_state=42)
kmeans.fit(X)
labels = kmeans.labels_
""", language="python")


st.markdown("""
- **`PCA()`**: Reduces the dimensionality of the data while preserving as much variance as possible.  
  *(Aids in visualizing high-dimensional data.)*
""")
st.code("""
from sklearn.decomposition import PCA

pca = PCA(n_components=2)
X_reduced = pca.fit_transform(X)
""", language="python")


st.markdown("""
- **`GridSearchCV()`**: Exhaustively searches over specified parameter values for an estimator.  
  *(Helps in finding the best model parameters.)*
""")
st.code("""
from sklearn.model_selection import GridSearchCV
from sklearn.ensemble import RandomForestClassifier

param_grid = {'n_estimators': [50, 100, 200], 'max_depth': [None, 10, 20]}
grid_search = GridSearchCV(RandomForestClassifier(), param_grid, cv=5)
grid_search.fit(X_train, y_train)
best_params = grid_search.best_params_
""", language="python")


st.markdown("""
- **`cross_val_score()`**: Evaluates a score by cross-validation.  
  *(Provides insights into model performance across different data splits.)*
""")
st.code("""
from sklearn.model_selection import cross_val_score
from sklearn.svm import SVC

model = SVC()
scores = cross_val_score(model, X, y, cv=5)
average_score = scores.mean()
""", language="python")


st.subheader("Statsmodels: The Statistical Powerhouse")
image_path_statmodels = "statmodels2.png"
st.image(image_path_statmodels, use_container_width=True)


st.markdown("""
Statsmodels is the go-to library for statistical modeling and analysis in Python. 
It excels in building detailed and interpretable models, making it a perfect complement to machine learning tools.
""")


st.markdown("""
- **`OLS()`**: Fits an Ordinary Least Squares regression model.  
  *(Best for analyzing relationships between variables.)*
""")
st.code("""
import statsmodels.api as sm

# Adding a constant for the intercept
X_const = sm.add_constant(X)

# Fitting an OLS regression model
model = sm.OLS(y, X_const)
results = model.fit()
print(results.summary())
""", language="python")


st.markdown("""
- **`GLM()`**: Fits Generalized Linear Models for flexible modeling.  
  *(Handles non-normal distributions like Poisson or Binomial.)*
""")
st.code("""
# Fitting a GLM
glm_model = sm.GLM(y, X_const, family=sm.families.Binomial())
glm_results = glm_model.fit()
print(glm_results.summary())
""", language="python")

# Time Series Analysis with ARIMA
st.markdown("""
- **`ARIMA()`**: Fits AutoRegressive Integrated Moving Average models.  
  *(Great for forecasting and analyzing time series data.)*
""")
st.code("""
from statsmodels.tsa.arima.model import ARIMA

# Fitting an ARIMA model
arima_model = ARIMA(time_series_data, order=(2, 1, 2))
arima_results = arima_model.fit()
print(arima_results.summary())
""", language="python")


st.markdown("""
- **`ttest_ind()`**: Performs a t-test for the means of two independent samples.  
  *(Perfect for hypothesis testing.)*
""")
st.code("""
from scipy.stats import ttest_ind

# Performing a t-test
stat, p_value = ttest_ind(sample1, sample2)
print(f"t-statistic: {stat}, p-value: {p_value}")
""", language="python")


st.markdown("""
- **`anova_lm()`**: Performs Analysis of Variance (ANOVA) on linear models.  
  *(Used for comparing multiple groups or models.)*
""")
st.code("""
from statsmodels.api import ols
from statsmodels.stats.anova import anova_lm

# Fitting a model and performing ANOVA
model = ols('dependent ~ C(group)', data=df).fit()
anova_results = anova_lm(model)
print(anova_results)
""", language="python")



st.subheader("TensorFlow: The Deep Learning Dynamo")
image_path_tensorflow = "tensor.png"
st.image(image_path_tensorflow, caption= "https://www.wired.com/2015/11/google-open-sources-its-artificial-intelligence-engine/",use_container_width=True)
st.warning("""
**Heads up!** PyTorch is a popular alternative to TensorFlow.  
- Use **TensorFlow** if you’re focused on production deployment, large-scale applications, or prefer a high-level API like Keras.
- Use **PyTorch** if you’re in a research environment, need flexibility, or prefer Pythonic code with dynamic computation graphs.
""")


st.markdown("""
- **`Sequential()`**: Constructs neural networks layer by layer, like assembling a high-tech sandwich.  
  *(Perfect for building deep learning models with all the fixings.)*
""")
st.code("""
import tensorflow as tf
from tensorflow.keras import Sequential
from tensorflow.keras.layers import Dense

# Build a deep learning model with layers stacked like a gourmet sandwich
model = Sequential([
    Dense(64, activation='relu', input_shape=(10,)),
    Dense(1, activation='sigmoid')
])
model.compile(optimizer='adam', loss='binary_crossentropy')
""", language="python")


st.markdown("""
- **`Functional API`**: Allows building complex models with shared layers or multiple inputs and outputs.  
  *(When the sandwich just isn’t enough, go gourmet!)*
""")
st.code("""
from tensorflow.keras import Model, Input
from tensorflow.keras.layers import Dense

# Create a functional model
inputs = Input(shape=(10,))
x = Dense(64, activation='relu')(inputs)
outputs = Dense(1, activation='sigmoid')(x)
functional_model = Model(inputs=inputs, outputs=outputs)
functional_model.compile(optimizer='adam', loss='binary_crossentropy')
""", language="python")


st.markdown("""
- **`callbacks`**: Automates training tasks like early stopping, saving models, or reducing learning rates.  
  *(Keeps your training on track without babysitting.)*
""")
st.code("""
from tensorflow.keras.callbacks import EarlyStopping

# Early stopping to prevent overfitting
early_stop = EarlyStopping(monitor='val_loss', patience=5)
model.fit(X_train, y_train, epochs=100, validation_split=0.2, callbacks=[early_stop])
""", language="python")


st.markdown("""
- **`TensorBoard`**: Visualizes training metrics, graphs, and more in a browser.  
  *(Your neural network’s personal dashboard.)*
""")
st.code("""
import tensorflow as tf
from tensorflow.keras.callbacks import TensorBoard

# Set up TensorBoard callback
tensorboard = TensorBoard(log_dir='./logs')
model.fit(X_train, y_train, epochs=10, callbacks=[tensorboard])
""", language="python")


st.markdown("""
- **`Pre-trained Models`**: Access state-of-the-art models like MobileNet or EfficientNet.  
  *(Why build from scratch when the pros have already done it?)*
""")
st.code("""
from tensorflow.keras.applications import MobileNetV2

# Load a pre-trained MobileNetV2 model
pretrained_model = MobileNetV2(weights='imagenet', include_top=False, input_shape=(224, 224, 3))
pretrained_model.summary()
""", language="python")

st.subheader("NLTK: The NLP Classic")
image_path_nltk = "nltk2.png"
st.image(image_path_nltk, use_container_width=True)

st.warning("""
**Heads up!** spaCy is a modern alternative to NLTK.  
- Use **NLTK** if you’re learning NLP concepts, experimenting with custom algorithms, or working on small-scale projects.
- Use **spaCy** if you need state-of-the-art pretrained models, speed, and production-ready tools for large-scale NLP tasks.
""")


st.markdown("""
NLTK (Natural Language Toolkit) is one of the earliest and most widely used libraries for Natural Language Processing (NLP). 
It’s an educational powerhouse, offering a range of tools and datasets for text analysis and linguistic tasks.
""")


st.markdown("""
- **Tokenization**: Splits text into words or sentences, enabling downstream analysis.  
  *(Helps break text into manageable pieces for NLP tasks.)*
""")
st.code("""
from nltk.tokenize import word_tokenize

text = "This is a test sentence."
tokens = word_tokenize(text)
print(tokens)
""", language="python")


st.markdown("""
- **Part-of-Speech (POS) Tagging**: Identifies grammatical tags for words, like nouns or verbs.  
  *(Useful for linguistic analysis and syntactic parsing.)*
""")
st.code("""
from nltk import pos_tag

tokens = word_tokenize("This is a test sentence.")
tags = pos_tag(tokens)
print(tags)
""", language="python")


st.markdown("""
- **Stop Words Removal**: Filters out common words like 'the' or 'and' that don’t carry meaningful information.  
  *(Helps focus on important words in text analysis.)*
""")
st.code("""
from nltk.corpus import stopwords

stop_words = set(stopwords.words("english"))
filtered_tokens = [word for word in tokens if word.lower() not in stop_words]
print(filtered_tokens)
""", language="python")


st.markdown("""
- **Stemming and Lemmatization**: Reduces words to their root form (e.g., 'running' to 'run').  
  *(Essential for normalizing text in NLP workflows.)*
""")
st.code("""
from nltk.stem import PorterStemmer
from nltk.stem import WordNetLemmatizer

stemmer = PorterStemmer()
lemmatizer = WordNetLemmatizer()

word = "running"
print("Stemmed:", stemmer.stem(word))
print("Lemmatized:", lemmatizer.lemmatize(word, pos='v'))
""", language="python")


st.subheader("XGBoost: The Gradient Boosting Powerhouse")
image_path_xgboost = "xgboost.png"
st.image(image_path_xgboost, use_container_width=True)

st.warning("""
**Heads up!** LightGBM is an alternative gradient boosting framework.  
- Use **XGBoost** if you’re working on smaller datasets or want a more widely recognized library in competitions like Kaggle.
- Use **LightGBM** if you’re handling large datasets, high-dimensional data, or need automatic handling of missing values.
""")

st.markdown("""
XGBoost (eXtreme Gradient Boosting) is a powerful machine learning library designed for structured data tasks. 
It’s widely used in data science competitions and excels in predictive accuracy and speed.
""")


st.markdown("""
- **Training a Model**: Quickly fit a gradient boosting model to structured data.  
  *(Handles missing values and categorical features with ease.)*
""")
st.code("""
from xgboost import XGBClassifier

# Training an XGBoost model
model = XGBClassifier()
model.fit(X_train, y_train)
predictions = model.predict(X_test)
""", language="python")


st.markdown("""
- **Feature Importance**: Identify the most influential features in your dataset.  
  *(Great for model interpretability and feature selection.)*
""")
st.code("""
from xgboost import plot_importance
import matplotlib.pyplot as plt

# Plotting feature importance
plot_importance(model)
plt.show()
""", language="python")


st.markdown("""
- **Hyperparameter Tuning**: Fine-tune model parameters like learning rate and tree depth for optimal performance.  
  *(Boosts model accuracy and reduces overfitting.)*
""")
st.code("""
from sklearn.model_selection import GridSearchCV

param_grid = {
    'learning_rate': [0.01, 0.1, 0.2],
    'max_depth': [3, 5, 7],
    'n_estimators': [50, 100, 200]
}

grid_search = GridSearchCV(XGBClassifier(), param_grid, cv=5)
grid_search.fit(X_train, y_train)
print("Best parameters:", grid_search.best_params_)
""", language="python")


st.markdown("""
- **Handling Missing Data**: XGBoost automatically handles missing values during training.  
  *(No need for imputation, making preprocessing simpler.)*
""")
st.code("""
# XGBoost can handle missing values (e.g., np.nan) natively
import numpy as np
X_train_with_missing = np.array([[np.nan, 2, 3], [4, 5, np.nan], [7, 8, 9]])
model.fit(X_train_with_missing, y_train)
""", language="python")



st.title("Essential Libaries for Data Engineers")
image_path_engineer = "de2.png"
st.image(image_path_engineer, use_column_width=True)
st.markdown("To avoid redundency, Pandas and Numpy are not listed here.")


st.subheader("SQLAlchemy: Your personal SQL translator")
image_path_sqlalchemy = "sqlalchemy.png"
st.image(image_path_sqlalchemy, use_container_width=True)

st.markdown("""
SQLAlchemy is a powerful library for interacting with databases using Python. 
It supports SQL queries and Object Relational Mapping (ORM), making database interaction smooth and Pythonic.
""")


st.markdown("""
- **Connecting to a Database**: Establishes a connection with your database for seamless data querying.  
  *(Supports multiple database engines like PostgreSQL, MySQL, and SQLite.)*
""")
st.code("""
from sqlalchemy import create_engine

# Connecting to a SQLite database
engine = create_engine('sqlite:///example.db')
connection = engine.connect()
""", language="python")


st.markdown("""
- **Using ORM**: Maps database tables to Python classes for intuitive interaction.  
  *(Great for abstracting SQL queries into Python code.)*
""")
st.code("""
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy import Column, Integer, String

Base = declarative_base()

class User(Base):
    __tablename__ = 'users'
    id = Column(Integer, primary_key=True)
    name = Column(String)

Base.metadata.create_all(engine)
""", language="python")


st.markdown("""
- **Querying with ORM**: Fetches and filters data using Pythonic syntax.
""")
st.code("""
from sqlalchemy.orm import sessionmaker

Session = sessionmaker(bind=engine)
session = Session()

# Adding a new user
new_user = User(name='Alice')
session.add(new_user)
session.commit()

# Querying users
users = session.query(User).filter_by(name='Alice').all()
print(users)
""", language="python")


st.markdown("""
- **Updating Data**: Modify existing records in the database.
""")
st.code("""
# Updating a user's name
user = session.query(User).filter_by(name='Alice').first()
user.name = 'Alicia'
session.commit()
""", language="python")


st.markdown("""
- **Deleting Data**: Remove records from the database with ease.
""")
st.code("""
# Deleting a user
user_to_delete = session.query(User).filter_by(name='Alicia').first()
session.delete(user_to_delete)
session.commit()
""", language="python")


st.subheader("Airflow: The Workflow Orchestrator")
image_path_airflow = "airflow2.png"
st.image(image_path_airflow, use_container_width=True)

st.markdown("""
Apache Airflow is a platform for orchestrating complex workflows. It schedules and monitors tasks, ensuring smooth ETL processes.
""")


st.markdown("""
- **Defining a DAG**: Create a Directed Acyclic Graph (DAG) to represent workflows.
""")
st.code("""
from airflow import DAG
from airflow.operators.dummy import DummyOperator
from datetime import datetime

with DAG('example_dag', start_date=datetime(2023, 1, 1), schedule_interval='@daily') as dag:
    start = DummyOperator(task_id='start')
    end = DummyOperator(task_id='end')
    start >> end
""", language="python")

image_path_dag = "dag.png"
st.image(image_path_dag, use_container_width=True)


st.markdown("""
- **Using PythonOperator**: Execute Python functions as tasks within a DAG.
""")
st.code("""
from airflow.operators.python import PythonOperator

def print_hello():
    print("Hello, Airflow!")

hello_task = PythonOperator(
    task_id='hello_task',
    python_callable=print_hello,
    dag=dag
)
""", language="python")


st.markdown("""
- **Setting Task Dependencies**: Manage the order of task execution in a DAG.
""")
st.code("""
# Setting task dependencies
start >> hello_task >> end
""", language="python")


st.markdown("""
- **Triggering DAGs**: Launch DAGs manually via the CLI or API.
""")
st.code("""
# Trigger a DAG from the CLI
# airflow dags trigger -d example_dag
""", language="bash")


st.subheader("PySpark: The Distributed Data Engine")
image_path_pyspark = "pyspark2.png"
st.image(image_path_pyspark, caption= "https://sparkbyexamples.com/pyspark/what-is-pyspark-and-who-uses-it/",use_container_width=True)

st.markdown("""
PySpark is the Python API for Apache Spark, enabling distributed data processing and large-scale computations.
""")


st.markdown("""
- **Creating a DataFrame**: Create distributed datasets for large-scale processing.
""")
st.code("""
from pyspark.sql import SparkSession

spark = SparkSession.builder.appName('example').getOrCreate()
data = [("Alice", 34), ("Bob", 23)]
columns = ["Name", "Age"]
df = spark.createDataFrame(data, columns)
df.show()
""", language="python")


st.markdown("""
- **Aggregations**: Perform operations like sum, average, or count on distributed data.
""")
st.code("""
df.groupBy("Name").avg("Age").show()
""", language="python")


st.markdown("""
- **Joins**: Combine data from multiple DataFrames.
""")
st.code("""
df2 = spark.createDataFrame([("Alice", "HR"), ("Bob", "IT")], ["Name", "Department"])
df.join(df2, on="Name", how="inner").show()
""", language="python")


st.markdown("""
- **Saving Data**: Write DataFrames to storage formats like Parquet or CSV.
""")
st.code("""
df.write.csv("output.csv")
""", language="python")


st.subheader("Dask: Parallel Computing Made Easy")
image_path_dask = "dask2.png"
st.image(image_path_dask, use_container_width=True)

st.markdown("""
Dask scales your Python workflows across multiple cores or clusters. It extends familiar libraries like Pandas, NumPy, and Scikit-learn to handle larger-than-memory computations.
""")


st.markdown("""
- **Dask DataFrame**: Scales Pandas-like operations to larger-than-memory datasets.  
  *(Useful for data too large to fit into memory.)*
""")
st.code("""
import dask.dataframe as dd

# Load a large dataset
df = dd.read_csv("large_dataset.csv")
print(df.head())
""", language="python")


st.markdown("""
- **Parallel Processing**: Run computations across multiple cores.  
  *(Speeds up operations for large datasets.)*
""")
st.code("""
# Compute a large operation in parallel
result = df.groupby("column").mean().compute()
print(result)
""", language="python")


st.markdown("""
- **Using the Dask Scheduler**: Manage distributed tasks on clusters.  
  *(Helps scale computations across machines.)*
""")
st.code("""
from dask.distributed import Client

# Connect to the Dask scheduler
client = Client()
print(client)
""", language="python")



st.subheader("FastAPI: The High-Performance API Framework")
image_path_fastapi = "fastapi.png"
st.image(image_path_fastapi, use_container_width=True)

st.markdown("""
FastAPI is a modern framework for building APIs with Python. It's fast, easy to use, and supports asynchronous programming for high performance.
""")


st.markdown("""
- **Creating a Basic API Endpoint**: Build a "Hello, World!" API in seconds.  
  *(Perfect for quick prototyping or full-scale APIs.)*
""")
st.code("""
from fastapi import FastAPI

app = FastAPI()

@app.get("/")
def read_root():
    return {"message": "Hello, FastAPI!"}
""", language="python")


st.markdown("""
- **Query Parameters**: Handle user inputs via URL parameters.  
  *(Great for filtering and dynamic data.)*
""")
st.code("""
@app.get("/items/")
def read_item(item_id: int, q: str = None):
    return {"item_id": item_id, "q": q}
""", language="python")


st.markdown("""
- **Asynchronous Support**: Leverage `async` functions for non-blocking operations.  
  *(Perfect for high-performance APIs.)*
""")
st.code("""
@app.get("/async-example/")
async def async_example():
    return {"message": "This endpoint is asynchronous!"}
""", language="python")


st.subheader("PyArrow: The Data Transport Specialist")
image_path_pyarrow = "pyarrow2.png"
st.image(image_path_pyarrow, use_container_width=True)

st.markdown("""
PyArrow enables fast data serialization, in-memory columnar data processing, and interoperability with other languages. 
It’s ideal for handling large datasets efficiently.
""")


st.markdown("""
- **Reading and Writing Parquet Files**: Efficiently handle columnar data formats for big data.  
  *(Optimized for speed and storage.)*
""")
st.code("""
import pyarrow.parquet as pq

# Write a Parquet file
pq.write_table(pyarrow.Table.from_pandas(df), "example.parquet")

# Read a Parquet file
table = pq.read_table("example.parquet")
print(table.to_pandas())
""", language="python")


st.markdown("""
- **Arrow Tables**: Work with in-memory columnar data.  
  *(Perfect for sharing data between Python and other systems.)*
""")
st.code("""
import pyarrow as pa

# Create an Arrow Table
table = pa.Table.from_pandas(df)
print(table.schema)
""", language="python")


st.markdown("""
- **Zero-Copy Data Sharing**: Transfer data between Python and other languages without duplication.  
  *(Speeds up workflows involving multiple systems.)*
""")
st.code("""
# Example of zero-copy sharing
arrow_array = pa.array([1, 2, 3])
print(arrow_array)
""", language="python")


st.subheader("Kafka: The Streaming Data Dynamo")
image_path_kafka = "kafka.png"
st.image(image_path_kafka,caption= "Franz Kafka",use_container_width=True)

st.markdown("""
Kafka is a distributed event-streaming platform designed for real-time data pipelines and analytics. It ensures fault-tolerant, high-throughput data streaming.
""")


st.markdown("""
- **Producing Messages**: Send data to Kafka topics for real-time streaming.  
  *(Enables scalable pipelines for data ingestion.)*
""")
st.code("""
from kafka import KafkaProducer

producer = KafkaProducer(bootstrap_servers='localhost:9092')
producer.send('example_topic', b'Hello, Kafka!')
producer.flush()
""", language="python")


st.markdown("""
- **Consuming Messages**: Retrieve data from Kafka topics for processing.  
  *(Perfect for real-time analytics.)*
""")
st.code("""
from kafka import KafkaConsumer

consumer = KafkaConsumer('example_topic', bootstrap_servers='localhost:9092')
for message in consumer:
    print(f"Received: {message.value}")
""", language="python")


st.markdown("""
- **Partitioning**: Distribute messages across multiple partitions for scalability.  
  *(Ensures high throughput and fault tolerance.)*
""")
st.code("""
# Produce messages with a specific partition
producer.send('example_topic', key=b'key1', value=b'value1', partition=0)
""", language="python")


st.title("Essential Libaries for Analytics Engineers/ML Engineers")
image_path_analytics = "Aeng.png"
st.image(image_path_analytics, use_container_width=True)

st.markdown("To avoid redundency, Pandas, Numpy, Scikit,Statsmodels, Tensorflow, Airflow, and Dask are not listed here.")

st.subheader("MLflow: The Model Lifecycle Manager")
image_path_mlflow = "mlflow.png"
st.image(image_path_mlflow, use_container_width=True)

st.markdown("""
MLflow is an open-source platform for managing the end-to-end lifecycle of machine learning models. It simplifies tracking experiments, packaging code, and deploying models into production.
""")


st.markdown("""
- **Experiment Tracking**: Log model parameters, metrics, and artifacts.  
  *(Helps compare model performance across runs.)*
""")
st.code("""
import mlflow

# Start a new MLflow run
with mlflow.start_run():
    mlflow.log_param("param1", 5)
    mlflow.log_metric("accuracy", 0.85)
    mlflow.log_artifact("model.pkl")
""", language="python")


st.markdown("""
- **Model Registry**: Organize and version your trained models.  
  *(Ensures reproducibility and ease of deployment.)*
""")
st.code("""
# Register a model
mlflow.register_model("runs:/example-run-id/model", "model_name")
""", language="python")


st.markdown("""
- **Deployment**: Deploy models to production with a single command.  
  *(Supports deployment to platforms like AWS SageMaker, Azure ML, or local servers.)*
""")
st.code("""
# Serve the model locally
mlflow.models.serve("model_name", port=1234)
""", language="python")


st.subheader("Hugging Face Transformers: The NLP Sorcerer")
image_path_huggingface = "huggingface.png"
st.image(image_path_huggingface, use_container_width=True)

st.markdown("""
Hugging Face Transformers is a library for state-of-the-art natural language processing (NLP). 
It provides pre-trained models for tasks like sentiment analysis, text generation, translation, and more.
""")


st.markdown("""
- **Sentiment Analysis**: Analyze the sentiment of text using a pre-trained model.  
  *(Quickly extract insights from textual data.)*
""")
st.code("""
from transformers import pipeline

# Load a sentiment analysis pipeline
sentiment_pipeline = pipeline("sentiment-analysis")
result = sentiment_pipeline("I love using Hugging Face Transformers!")
print(result)
""", language="python")


st.markdown("""
- **Text Generation**: Generate text using a language model like GPT.  
  *(Great for creative applications or conversational AI.)*
""")
st.code("""
generator = pipeline("text-generation", model="gpt2")
result = generator("Once upon a time", max_length=50, num_return_sequences=1)
print(result)
""", language="python")


st.markdown("""
- **Fine-Tuning**: Customize pre-trained models for your specific dataset.  
  *(Perfect for adapting models to domain-specific tasks.)*
""")
st.code("""
from transformers import AutoModelForSequenceClassification, AutoTokenizer
from transformers import TrainingArguments, Trainer

model = AutoModelForSequenceClassification.from_pretrained("distilbert-base-uncased")
tokenizer = AutoTokenizer.from_pretrained("distilbert-base-uncased")

# Tokenize dataset and fine-tune the model using Trainer API
# Example assumes `train_dataset` and `test_dataset` are preprocessed
training_args = TrainingArguments(output_dir="./results", num_train_epochs=3, per_device_train_batch_size=8)
trainer = Trainer(model=model, args=training_args, train_dataset=train_dataset, eval_dataset=test_dataset)
trainer.train()
""", language="python")


st.subheader("CatBoost: The Categorical Data Champion")
image_path_catboost = "catboost2.png"
st.image(image_path_catboost,caption="https://towardsdatascience.com/why-you-should-learn-catboost-now-390fb3895f76", use_container_width=True)

st.markdown("""
CatBoost is a gradient boosting library that handles categorical data natively. 
It’s highly efficient and achieves excellent performance with minimal preprocessing.
""")


st.markdown("""
- **Training a Model**: Train a gradient boosting model for classification or regression tasks.  
  *(Handles categorical features automatically.)*
""")
st.code("""
from catboost import CatBoostClassifier

# Initialize and train the model
model = CatBoostClassifier(iterations=100, depth=10, learning_rate=0.1, verbose=0)
model.fit(X_train, y_train, cat_features=[0, 1])  # Specify categorical feature indices
""", language="python")


st.markdown("""
- **Feature Importance**: Identify which features contribute most to predictions.  
  *(Helps in model interpretability.)*
""")
st.code("""
# Get feature importance
importance = model.get_feature_importance()
print(importance)
""", language="python")


st.markdown("""
- **Cross-Validation**: Evaluate model performance using cross-validation.  
  *(Provides robust insights into model accuracy.)*
""")
st.code("""
from catboost import cv, Pool

# Perform cross-validation
pool = Pool(X_train, y_train, cat_features=[0, 1])
cv_results = cv(pool, params={"iterations": 50, "depth": 5, "learning_rate": 0.1}, fold_count=5)
print(cv_results)
""", language="python")


st.subheader("ONNX: The Interoperable Model Format")
image_path_onnx = "onnx2.png"
st.image(image_path_onnx, use_container_width=True)

st.markdown("""
ONNX (Open Neural Network Exchange) provides a universal format for machine learning models. 
It simplifies model conversion, optimization, and inference across different tools or platforms.
""")

# Exporting a Model to ONNX
st.markdown("""
- **Exporting Models**: Convert models to ONNX format for interoperability.  
  *(Works with frameworks like PyTorch, TensorFlow, and Scikit-learn.)*
""")
st.code("""
import torch
import onnx

# Export a PyTorch model to ONNX
torch_model = MyModel()
dummy_input = torch.randn(1, 3, 224, 224)
torch.onnx.export(torch_model, dummy_input, "model.onnx")
""", language="python")


st.markdown("""
- **Inference with ONNX Runtime**: Run models efficiently using ONNX Runtime.  
  *(Optimized for speed and portability.)*
""")
st.code("""
import onnxruntime as ort

# Load ONNX model and perform inference
session = ort.InferenceSession("model.onnx")
inputs = {"input": input_data}
outputs = session.run(None, inputs)
print(outputs)
""", language="python")

# Model Conversion
st.markdown("""
- **Model Conversion**: Convert models from frameworks like TensorFlow or Scikit-learn.  
  *(Ensures compatibility across tools.)*
""")
st.code("""
from skl2onnx import convert_sklearn
from skl2onnx.common.data_types import FloatTensorType

# Convert a Scikit-learn model to ONNX
initial_type = [("float_input", FloatTensorType([None, X.shape[1]]))]
onnx_model = convert_sklearn(sklearn_model, initial_types=initial_type)
with open("model.onnx", "wb") as f:
    f.write(onnx_model.SerializeToString())
""", language="python")



# Ensure the required datasets exist
def create_sample_datasets():
    datasets = {
        "sales_data.csv": """Item,Price,Quantity\nWidget,10,5\nGadget,15,3\nDoohickey,7,10""",
        "employees.csv": """Name,Department,Salary\nAlice,HR,60000\nBob,Engineering,75000\nCharlie,Marketing,50000""",
        "weather_data.csv": """Date,Temperature,Rainfall\n2023-01-01,32,0.1\n2023-01-02,30,0.2\n2023-01-03,28,0.0""",
    }

    for filename, content in datasets.items():
        if not os.path.exists(filename):
            with open(filename, "w") as f:
                f.write(content)

# Create datasets if they don't exist
create_sample_datasets()

# Streamlit app starts here
st.title("Python Sandbox for Data Exploration")

# Sidebar for sample datasets
st.sidebar.header("Sample Datasets")
datasets = {
    "Sales Data": "sales_data.csv",
    "Employee Records": "employees.csv",
    "Weather Data": "weather_data.csv",
}

selected_dataset = st.sidebar.selectbox("Choose a dataset to explore:", list(datasets.keys()))

# Load the selected dataset
df = pd.read_csv(datasets[selected_dataset])
st.write(f"### Loaded Dataset: {selected_dataset}")
st.dataframe(df)

# Run Python Code Section
st.subheader("Run Python Code")
code = st.text_area("Write your Python code here and press 'Run':", value="""
# Example:
df['New Column'] = df['Price'] * df['Quantity']
df.head()
""")

if st.button("Run"):
    try:
        # Restrict execution to a safe namespace
        local_vars = {"df": df}
        exec(code, {"pd": pd, "np": np}, local_vars)  # Allow only necessary libraries
        st.write("### Resulting DataFrame:")
        st.dataframe(local_vars["df"])

        # Allow the user to download the modified dataset
        csv = local_vars["df"].to_csv(index=False).encode("utf-8")
        st.download_button(
            label="Download CSV",
            data=csv,
            file_name="modified_dataset.csv",
            mime="text/csv",
        )
    except Exception as e:
        st.error(f"Error: {e}")
        st.error("Please ensure your code is valid and doesn’t include risky operations like file deletions.")

# File Uploader Section
uploaded_file = st.file_uploader("Upload your CSV file:", type="csv")
if uploaded_file:
    try:
        # Preview the uploaded file
        preview_df = pd.read_csv(uploaded_file, nrows=5)
        st.write("### File Preview (First 5 Rows):")
        st.dataframe(preview_df)

        # Option to confirm loading
        if st.button("Load Uploaded Dataset"):
            df = pd.read_csv(uploaded_file)
            st.success("Dataset loaded successfully!")
            st.write("### Loaded Dataset from Upload:")
            st.dataframe(df)
    except Exception as e:
        st.error(f"Error loading file: {e}")

# Visualization Section
st.subheader("Quick Visualizations")
if st.checkbox("Bar Chart"):
    x_col = st.selectbox("Select X-axis column:", df.columns)
    y_col = st.selectbox("Select Y-axis column:", df.columns)
    st.bar_chart(df.groupby(x_col)[y_col].sum())

if st.checkbox("Line Chart"):
    line_col = st.selectbox("Select column for Line Chart:", df.columns)
    st.line_chart(df[line_col])

if st.checkbox("Scatter Plot"):
    x_col = st.selectbox("Select X-axis for Scatter Plot:", df.columns)
    y_col = st.selectbox("Select Y-axis for Scatter Plot:", df.columns)
    st.pyplot(df.plot.scatter(x=x_col, y=y_col).get_figure())
