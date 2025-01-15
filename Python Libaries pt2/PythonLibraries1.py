
import streamlit as st


import pandas as pd
import numpy as np


import matplotlib.pyplot as plt
import seaborn as sns
import plotly.express as px


from sklearn.linear_model import LinearRegression
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_squared_error


from sqlalchemy import create_engine # type: ignore


import requests

st.title("Essential Python Libraries for Analysts, Scientists, and Engineers")
image_path_python = "pythoncloud.jpg"
st.image(image_path_python, use_container_width=True)
st.write("Welcome to the Essential Python Libraries for Analysts, Scientists, and Engineers! This is a Streamlit app that showcases the most important Python libraries for data analysis, visualization, and machine learning. The libraries that will be covered in this app are: Pandas, NumPy, Matplotlib, Seaborn, Plotly, and Scikit-Learn. Each library will have its own section where we will explore the most important features and functions. Let's get started!")


st.title("Essential Libaries for Data Analysts (If your up for it)")
image_path_analyst = "data-analyst1.jpg"
st.image(image_path_analyst, use_container_width=True)


st.subheader("1. Pandas: The Data Manipulation Powerhouse")
image_path_pandas = "Pandas.png"
st.image(image_path_pandas, use_container_width=True)
st.warning(""" Use Case: Sales Data Analysis (aka ‘Make It Pretty for the Executive Team’)""")
st.markdown(""" You’ve got sales data that looks like it was written in a hurry by a caffeinated jack russell. Pandas helps you clean it up, calculate revenues, and make it look like you spent hours on it (when it really took 15 minutes).
""")


st.markdown("""
- **`read_csv()`**: Reads your messy CSV file and pretends there is clean data inside.""")
st.code("""
import pandas as pd

# Read a CSV file
data = pd.read_csv("sales_data.csv")
print(data.head())
""", language="python")

st.markdown("""
- **`groupby()`**: Groups data like a champ.""")
st.code("""
# Group data by a column and calculate sum
grouped_data = data.groupby("Product")["Revenue"].sum()
print(grouped_data)
""", language="python")


st.markdown("""
- **`merge()`**: Matches data like a Tinder date—sometimes perfect, sometimes… eh.
""")
st.code("""
# Merge two DataFrames on a common column
df1 = pd.DataFrame({"ID": [1, 2], "Name": ["Alice", "Bob"]})
df2 = pd.DataFrame({"ID": [1, 2], "Score": [85, 90]})
merged_data = pd.merge(df1, df2, on="ID")
print(merged_data)
""", language="python")


st.markdown("""
- **`fillna()`**: Fills in the blanks.""")  

st.code("""
# Fill missing values with a default value
data["Revenue"] = data["Revenue"].fillna(0)
print(data.head())
""", language="python")


st.markdown("""
- **`describe()`**: Summarizes your data so you can sound smarter than you feel.""") 

st.code("""
# Get summary statistics
summary = data.describe()
print(summary)
""", language="python")



st.markdown("""
- **`pivot_table()`**: Creates fancy pivot tables that would make Excel jealous.""") 

st.code("""
# Create a pivot table
pivot = data.pivot_table(values="Revenue", index="Product", columns="Region", aggfunc="sum")
print(pivot)
""", language="python")


st.markdown("""
- **`dropna()`**: Kicks out rows or columns with missing data, like a bouncer at a nightclub.""")

st.code("""
# Drop rows with missing values
cleaned_data = data.dropna()
print(cleaned_data.head())
""", language="python")


st.markdown("""
- **`sort_values()`**: Sorts your data by column values because who doesn’t like order?""")

st.code("""
# Sort data by the 'Revenue' column in descending order
sorted_data = data.sort_values(by="Revenue", ascending=False)
print(sorted_data.head())
""", language="python")

st.markdown("""
- **`iloc[]` and `loc[]`**: Helps you find that one needle in the data haystack.""")

st.code("""
# iloc example: Get the first 5 rows
subset_iloc = data.iloc[:5]

# loc example: Get rows where 'Revenue' > 100
subset_loc = data.loc[data["Revenue"] > 100]
print(subset_iloc)
print(subset_loc)
""", language="python")


st.markdown("""
- **`apply()`**: Lets you apply your own logic to rows or columns—perfect for when Pandas doesn’t already do it for you.""")

st.code("""
# Apply a custom function to a column
data["Discounted_Price"] = data["Price"].apply(lambda x: x * 0.9)
print(data.head())
""", language="python")


st.markdown("""
- **`astype()`**: Converts data types, because sometimes your numbers come in disguised as text.""")

st.code("""
# Convert a column to float
data["Revenue"] = data["Revenue"].astype(float)
print(data.dtypes)
""", language="python")


st.markdown("""
- **`value_counts()`**: Counts how many times each value appears in a column, so you can spot trends or weird anomalies.""")

st.code("""
# Count the occurrences of each product
product_counts = data["Product"].value_counts()
print(product_counts)
""", language="python")

st.markdown("""
- **`concat()`**: Sticks multiple DataFrames together like a glue for data.""")

st.code("""
# Concatenate two DataFrames
df1 = pd.DataFrame({"ID": [1, 2], "Name": ["Alice", "Bob"]})
df2 = pd.DataFrame({"ID": [3, 4], "Name": ["Charlie", "Diana"]})
combined_data = pd.concat([df1, df2])
print(combined_data)
""", language="python")


st.markdown("""
- **`drop_duplicates()`**: Removes duplicate rows because redundancy in this context sucks.Why? Duplicate rows in your dataset can skew analysis, inflate metrics, and lead to incorrect insights. """)

st.code("""
# Remove duplicate rows
unique_data = data.drop_duplicates()
print(unique_data)
""", language="python")


st.markdown("""
- **`isnull()`**: Detects missing data faster than your boss finds typos in your reporting.""")

st.code("""
# Check for missing values
missing_data = data.isnull()
print(missing_data)
""", language="python")


st.markdown("""
- **`to_excel()`**: Exports your cleaned-up masterpiece back to Excel for the spreadsheet fans in your office.
""")


st.code("""
# Save the DataFrame to an Excel file
data.to_excel("cleaned_data.xlsx", index=False)
print("Data exported to Excel!")
""", language="python")


st.subheader(" NumPy: Number Cruncher")
image_path_pandas = "Numpy.png"
st.image(image_path_pandas, use_container_width=True)


st.markdown("""
- **`array()`**: Creates arrays that don’t complain like lists.  
  *(An array is a powerful, grid-like data structure provided by NumPy. Why it doesn’t complain: Arrays offer uniformity, speed, and efficient math operations!)*
""")

st.code("""
import numpy as np

# Create an array
arr = np.array([1, 2, 3, 4])
print(arr)
""", language="python")

st.markdown("""
- **`mean()`**: Finds the average, because math.
""")
st.code("""
# Calculate the mean
mean_value = np.mean(arr)
print(mean_value)
""", language="python")

st.markdown("""
- **`sum()`**: Adds things up faster than a calculator.
""")
st.code("""
# Compute the sum
total = np.sum(arr)
print(total)
""", language="python")

st.markdown("""
- **`reshape()`**: Reshapes arrays like a pro.
""")
st.code("""
# Reshape a 1D array into a 2D array
reshaped = arr.reshape(2, 2)
print(reshaped)
""", language="python")

st.markdown("""
- **`random()`**: Generates random numbers because why not(Be Careful, and Intelligent)?
""")
st.code("""
# Generate random numbers
random_numbers = np.random.rand(3)
print(random_numbers)
""", language="python")


st.subheader("Matplotlib: Paint Data Like Picasso")
image_path_matplotlib = "truematplot.png"
st.image(image_path_matplotlib, use_container_width=True)


st.markdown("""
- **`plot()`**: Connects the dots with lines like an old-school Etch A Sketch.  
  *(Use this for simple line charts, great for trends or continuous data.)*
""")
st.code("""
import matplotlib.pyplot as plt

# Create a line plot
x = [1, 2, 3, 4]
y = [10, 20, 25, 30]
plt.plot(x, y)
plt.title("Line Plot Example")
plt.show()
""", language="python")

st.markdown("""
- **`scatter()`**: Puts dots on a graph, perfect for showing relationships.  
  *(Use this for visualizing correlations or clusters.)*
""")
st.code("""
# Create a scatter plot
plt.scatter(x, y)
plt.title("Scatter Plot Example")
plt.show()
""", language="python")

image_path_graph = "Matplotgraph.png"
st.image(image_path_graph,caption="https://coding-blocks.github.io/DS-NOTES/1.2%20Introduction%20to%20Matplotlib.html", use_container_width=True)




st.subheader("Seaborn: Stylish Data Visualizations")
image_path_seaborn = "seaborn.png"
st.image(image_path_seaborn, use_container_width=True)


st.markdown("""
- **`heatmap()`**: Makes your data hot (or cold).  
  *(Use this to visualize correlations, null values, or matrix data.)*
""")
st.code("""
import seaborn as sns
import numpy as np

# Create a heatmap
data = np.random.rand(5, 5)
sns.heatmap(data, annot=True)
plt.title("Heatmap Example")
plt.show()
""", language="python")

st.markdown("""
- **`boxplot()`**: Puts your data in a box (literally) to show ranges and outliers.  
  *(Great for showing distributions and spotting anomalies.)*
""")
st.code("""
# Create a boxplot
sns.boxplot(data=[7, 8, 8, 9, 10, 10, 11])
plt.title("Boxplot Example")
plt.show()
""", language="python")

image_path_seaborngraph = "seaborngraph.png"
st.image(
    image_path_seaborngraph, 
    caption="[Seaborn Function Overview](https://seaborn.pydata.org/tutorial/function_overview.html)", 
    use_container_width=True
)


st.subheader("openpyxl: The Excel Whisperer")
image_path_openpyxl = "xyl.png"
st.image(image_path_openpyxl, use_container_width=True)


st.markdown("""
- **`load_workbook()`**: Opens an Excel file, no double-clicking required.  
  *(Use this to edit existing spreadsheets programmatically.)*
""")
st.code("""
from openpyxl import load_workbook

# Load an existing Excel workbook
workbook = load_workbook("example.xlsx")
sheet = workbook.active
print(sheet.title)
""", language="python")

st.markdown("""
- **`create_sheet()`**: Adds a shiny new sheet to your workbook.  
  *(Perfect for organizing data into multiple tabs.)*
""")
st.code("""
# Create a new sheet
workbook.create_sheet(title="NewSheet")
workbook.save("example.xlsx")
""", language="python")



st.title("Essential Libaries for Data Scientists")
image_path_science = "/c/Users/Sloan/Documents/Github/Streamlit_Series/DS.png"
st.image(image_path_science, use_column_width=True)

st.subheader("Pandas")

st.subheader("NumPy")

st.subheader("scikit-learn & Statsmodels & keras")

st.subheader("TensorFlow & PyTorch ")

st.subheader("NLTK")


st.title("Essential Libaries for Analytics Engineers/ML Engineers")
image_path_analytics = "/c/Users/Sloan/Documents/Github/Streamlit_Series/Aeng.png"
st.image(image_path_analytics, use_column_width=True)

st.markdown("To avoid redundency, Pandas and Numpy are not listed here.")

st.subheader("scikit-learn & Statsmodels & keras")

st.subheader("TensorFlow & PyTorch ")

st.subheader("Airflow")

st.subheader("Dask")

st.subheader("MLflow")


st.title("Essential Libaries for Data Engineers")
image_path_engineer = "/c/Users/Sloan/Documents/Github/Streamlit_Series/Engineers.jpg"
st.image(image_path_engineer, use_column_width=True)

st.markdown("To avoid redundency, Pandas and Numpy are not listed here.")

st.subheader("SQLAlchemy")

st.subheader("Airflow")

st.subheader("PySpark")

st.subheader("Dask")

st.subheader("FastAPI")

st.subheader("Pyarrow")

import pandas as pd
import streamlit as st


st.title("Python Sandbox for Data Exploration")


st.sidebar.header("📂 Sample Datasets")
datasets = {
    "Sales Data": "sales_data.csv",
    "Employee Records": "employees.csv",
    "Weather Data": "weather_data.csv",
}

selected_dataset = st.sidebar.selectbox("Choose a dataset to explore:", list(datasets.keys()))


df = pd.read_csv(datasets[selected_dataset])
st.write(f"### Loaded Dataset: {selected_dataset}")
st.dataframe(df)



st.subheader("Run Python Code")
code = st.text_area("Write your Python code here and press 'Run':", value="""
# Example:
df['New Column'] = df['Price'] * df['Quantity']
df.head()
""")


if st.button("Run"):
    try:
        local_vars = {"df": df}
        exec(code, {"pd": pd}, local_vars)

     
        st.write("### Resulting DataFrame:")
        st.dataframe(local_vars["df"])
    except Exception as e:
        st.error(f"Error: {e}")


uploaded_file = st.file_uploader("Upload your CSV file:", type="csv")
if uploaded_file:
    df = pd.read_csv(uploaded_file)
